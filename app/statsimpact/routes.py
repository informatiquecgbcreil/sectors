from __future__ import annotations

from datetime import date
from io import BytesIO, StringIO
import csv

import os

from flask import Blueprint, abort, render_template, request, redirect, url_for, flash, current_app, send_file, Response
from flask_login import login_required, current_user
from app.rbac import can

from sqlalchemy import func
from sqlalchemy.orm import joinedload

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.extensions import db

from app.models import (
    AtelierActivite,
    Participant,
    Quartier,
    PresenceActivite,
    SessionActivite,
    Projet,
    Competence,
    Evaluation,
    Referentiel,
    Objectif,
)

from app.activite.services.docx_utils import generate_participant_bilan_pdf

from .occupancy import compute_occupancy_stats

from .engine import (
    compute_volume_activity_stats,
    compute_participation_frequency_stats,
    compute_transversalite_stats,
    compute_demography_stats,
    compute_participants_stats,
    compute_magatomatique,
    normalize_filters,
    _apply_common_filters,
    _session_date_expr,
    _session_duration_minutes,
)

bp = Blueprint("statsimpact", __name__, url_prefix="")


CSV_FIELD_GROUPS = [
    {
        "label": "Participant",
        "fields": [
            {"key": "participant_id", "label": "ID participant"},
            {"key": "participant_nom", "label": "Nom"},
            {"key": "participant_prenom", "label": "Prénom"},
            {"key": "participant_email", "label": "Email"},
            {"key": "participant_telephone", "label": "Téléphone"},
            {"key": "participant_ville", "label": "Ville"},
            {"key": "participant_quartier", "label": "Quartier"},
            {"key": "participant_genre", "label": "Genre"},
            {"key": "participant_type_public", "label": "Type public"},
            {"key": "participant_date_naissance", "label": "Date naissance"},
        ],
    },
    {
        "label": "Session",
        "fields": [
            {"key": "session_id", "label": "ID session"},
            {"key": "session_date", "label": "Date session"},
            {"key": "session_type", "label": "Type session"},
            {"key": "session_statut", "label": "Statut session"},
            {"key": "session_heure_debut", "label": "Heure début"},
            {"key": "session_heure_fin", "label": "Heure fin"},
            {"key": "session_duree_minutes", "label": "Durée (minutes)"},
        ],
    },
    {
        "label": "Atelier",
        "fields": [
            {"key": "atelier_id", "label": "ID atelier"},
            {"key": "atelier_nom", "label": "Nom atelier"},
            {"key": "atelier_secteur", "label": "Secteur atelier"},
            {"key": "atelier_type", "label": "Type atelier"},
        ],
    },
    {
        "label": "Présence",
        "fields": [
            {"key": "presence_id", "label": "ID présence"},
            {"key": "presence_motif", "label": "Motif"},
            {"key": "presence_motif_autre", "label": "Motif autre"},
            {"key": "presence_created_at", "label": "Date d'émargement"},
        ],
    },
]

CSV_DEFAULT_FIELDS = [
    "participant_nom",
    "participant_prenom",
    "atelier_nom",
    "atelier_secteur",
    "session_date",
    "session_type",
]


def _fmt_date(value) -> str:
    return value.strftime("%Y-%m-%d") if value else ""


def _fmt_datetime(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


CSV_FIELD_MAP = {
    "participant_id": {"label": "ID participant", "getter": lambda ctx: ctx["participant"].id},
    "participant_nom": {"label": "Nom", "getter": lambda ctx: ctx["participant"].nom or ""},
    "participant_prenom": {"label": "Prénom", "getter": lambda ctx: ctx["participant"].prenom or ""},
    "participant_email": {"label": "Email", "getter": lambda ctx: ctx["participant"].email or ""},
    "participant_telephone": {"label": "Téléphone", "getter": lambda ctx: ctx["participant"].telephone or ""},
    "participant_ville": {"label": "Ville", "getter": lambda ctx: ctx["participant"].ville or ""},
    "participant_quartier": {
        "label": "Quartier",
        "getter": lambda ctx: ctx["quartier"].nom if ctx["quartier"] else "",
    },
    "participant_genre": {"label": "Genre", "getter": lambda ctx: ctx["participant"].genre or ""},
    "participant_type_public": {"label": "Type public", "getter": lambda ctx: ctx["participant"].type_public or ""},
    "participant_date_naissance": {
        "label": "Date naissance",
        "getter": lambda ctx: _fmt_date(ctx["participant"].date_naissance),
    },
    "session_id": {"label": "ID session", "getter": lambda ctx: ctx["session"].id},
    "session_date": {
        "label": "Date session",
        "getter": lambda ctx: _fmt_date(ctx["session"].rdv_date or ctx["session"].date_session),
    },
    "session_type": {"label": "Type session", "getter": lambda ctx: ctx["session"].session_type or ""},
    "session_statut": {"label": "Statut session", "getter": lambda ctx: ctx["session"].statut or ""},
    "session_heure_debut": {
        "label": "Heure début",
        "getter": lambda ctx: ctx["session"].rdv_debut or ctx["session"].heure_debut or "",
    },
    "session_heure_fin": {
        "label": "Heure fin",
        "getter": lambda ctx: ctx["session"].rdv_fin or ctx["session"].heure_fin or "",
    },
    "session_duree_minutes": {
        "label": "Durée (minutes)",
        "getter": lambda ctx: ctx["session"].duree_minutes or "",
    },
    "atelier_id": {"label": "ID atelier", "getter": lambda ctx: ctx["atelier"].id},
    "atelier_nom": {"label": "Nom atelier", "getter": lambda ctx: ctx["atelier"].nom or ""},
    "atelier_secteur": {"label": "Secteur atelier", "getter": lambda ctx: ctx["atelier"].secteur or ""},
    "atelier_type": {"label": "Type atelier", "getter": lambda ctx: ctx["atelier"].type_atelier or ""},
    "presence_id": {"label": "ID présence", "getter": lambda ctx: ctx["presence"].id},
    "presence_motif": {"label": "Motif", "getter": lambda ctx: ctx["presence"].motif or ""},
    "presence_motif_autre": {"label": "Motif autre", "getter": lambda ctx: ctx["presence"].motif_autre or ""},
    "presence_created_at": {
        "label": "Date d'émargement",
        "getter": lambda ctx: _fmt_datetime(ctx["presence"].created_at),
    },
}


def _can_view() -> bool:
    return can("statsimpact:view") or can("statsimpact:view_all")


def _safe_sheet_title(name: str, fallback: str = "Atelier") -> str:
    """Openpyxl: max 31 chars, no [ ] * ? / \\ etc."""
    if not name:
        name = fallback
    bad = set('[]:*?/\\')
    cleaned = "".join(c for c in name if c not in bad).strip()
    cleaned = cleaned[:31] if cleaned else fallback
    return cleaned


def _pedago_scope_secteur() -> str | None:
    if can("scope:all_secteurs"):
        return None
    return (getattr(current_user, "secteur_assigne", None) or "").strip() or None


def _build_bilan_rows(participant: Participant) -> list[dict]:
    eval_rows = (
        db.session.query(
            Evaluation,
            Competence,
            Referentiel,
            SessionActivite,
            AtelierActivite,
        )
        .join(Competence, Evaluation.competence_id == Competence.id)
        .join(Referentiel, Competence.referentiel_id == Referentiel.id)
        .outerjoin(SessionActivite, Evaluation.session_id == SessionActivite.id)
        .outerjoin(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
        .filter(Evaluation.participant_id == participant.id, Evaluation.etat == 2)
        .order_by(Referentiel.nom.asc(), Competence.code.asc(), Evaluation.date_evaluation.asc())
        .all()
    )

    rows: list[dict] = []
    for eval_obj, comp, ref, session, atelier in eval_rows:
        date_label = eval_obj.date_evaluation.strftime("%d/%m/%Y") if eval_obj.date_evaluation else ""
        atelier_label = atelier.nom if atelier else ""
        rows.append(
            {
                "referentiel": ref.nom,
                "competence": f"{comp.code} · {comp.nom}",
                "date": date_label,
                "atelier": atelier_label,
            }
        )
    return rows


def _participants_success_rate(session_id: int, competences: list[Competence]) -> dict:
    if not competences:
        return {"total": 0, "success": 0, "ratio": 0}
    presences = PresenceActivite.query.filter_by(session_id=session_id).all()
    total = len(presences)
    if total == 0:
        return {"total": 0, "success": 0, "ratio": 0}
    comp_ids = [c.id for c in competences]
    success_count = 0
    for pr in presences:
        evals = (
            Evaluation.query.filter(
                Evaluation.session_id == session_id,
                Evaluation.participant_id == pr.participant_id,
                Evaluation.competence_id.in_(comp_ids),
                Evaluation.etat >= 2,
            )
            .distinct()
            .count()
        )
        if evals == len(comp_ids):
            success_count += 1
    ratio = (success_count / total * 100) if total else 0
    return {"total": total, "success": success_count, "ratio": ratio}


def _objective_success(obj: Objectif) -> dict:
    if obj.type == "operationnel" and obj.session_id:
        stats = _participants_success_rate(obj.session_id, obj.competences)
        validated = stats["ratio"] >= (obj.seuil_validation or 0)
        return {"ratio": stats["ratio"], "validated": validated, "total": stats["total"], "success": stats["success"]}

    enfants = obj.enfants or []
    if not enfants:
        return {"ratio": 0, "validated": False, "total": 0, "success": 0}
    results = [ _objective_success(child) for child in enfants ]
    total = len(results)
    success = sum(1 for r in results if r["validated"])
    ratio = (success / total * 100) if total else 0
    validated = ratio >= (obj.seuil_validation or 0)
    return {"ratio": ratio, "validated": validated, "total": total, "success": success}


def _query_presence_export(flt, participant_q: str | None = None):
    query = (
        db.session.query(
            PresenceActivite,
            Participant,
            SessionActivite,
            AtelierActivite,
            Quartier,
        )
        .join(Participant, PresenceActivite.participant_id == Participant.id)
        .join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
        .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
        .outerjoin(Quartier, Participant.quartier_id == Quartier.id)
    )
    query = _apply_common_filters(query, flt)

    if participant_q:
        like = f"%{participant_q.lower()}%"
        query = query.filter(
            func.lower(func.coalesce(Participant.nom, "")).like(like)
            | func.lower(func.coalesce(Participant.prenom, "")).like(like)
        )

    query = query.order_by(_session_date_expr().asc(), Participant.nom.asc(), Participant.prenom.asc())
    return query


@bp.route("/stats/pedagogie", methods=["GET"])
@login_required
def stats_pedagogie():
    if not _can_view():
        abort(403)

    secteur = _pedago_scope_secteur()

    projets_q = Projet.query
    ateliers_q = AtelierActivite.query.filter(AtelierActivite.is_deleted.is_(False))
    if secteur:
        projets_q = projets_q.filter(Projet.secteur == secteur)
        ateliers_q = ateliers_q.filter(AtelierActivite.secteur == secteur)

    projets = projets_q.order_by(Projet.secteur.asc(), Projet.nom.asc()).all()
    ateliers = ateliers_q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()
    participants = Participant.query.order_by(Participant.nom.asc(), Participant.prenom.asc()).all()

    projet_id = request.args.get("projet_id", type=int)
    atelier_id = request.args.get("atelier_id", type=int)
    participant_id = request.args.get("participant_id", type=int)

    projet = Projet.query.get(projet_id) if projet_id else None
    if projet and secteur and projet.secteur != secteur:
        abort(403)

    atelier = AtelierActivite.query.get(atelier_id) if atelier_id else None
    if atelier and secteur and atelier.secteur != secteur:
        abort(403)

    participant = Participant.query.get(participant_id) if participant_id else None

    projet_objectifs = []
    if projet:
        objectifs = Objectif.query.filter_by(projet_id=projet.id, type="general").order_by(Objectif.created_at.asc()).all()
        for obj in objectifs:
            stats = _objective_success(obj)
            projet_objectifs.append({"objectif": obj, **stats})

    atelier_stats = {}
    if atelier:
        objectifs = Objectif.query.filter_by(atelier_id=atelier.id, type="specifique").order_by(Objectif.created_at.asc()).all()
        objectifs_stats = []
        for obj in objectifs:
            stats = _objective_success(obj)
            objectifs_stats.append({"objectif": obj, **stats})
        atelier_stats = {"objectifs": objectifs_stats}

    participant_groups = []
    bilan_rows = []
    if participant:
        bilan_rows = _build_bilan_rows(participant)
        grouped: dict[str, list[dict]] = {}
        for row in bilan_rows:
            grouped.setdefault(row["referentiel"], []).append(row)
        participant_groups = [
            {"referentiel": ref, "rows": rows} for ref, rows in grouped.items()
        ]

    return render_template(
        "stats_pedagogie.html",
        projets=projets,
        ateliers=ateliers,
        participants=participants,
        projet=projet,
        atelier=atelier,
        participant=participant,
        projet_objectifs=projet_objectifs,
        atelier_stats=atelier_stats,
        participant_groups=participant_groups,
    )


@bp.route("/stats/pedagogie/participant/<int:participant_id>/bilan", methods=["GET"])
@login_required
def stats_pedagogie_bilan(participant_id: int):
    if not _can_view():
        abort(403)
    participant = Participant.query.get_or_404(participant_id)
    rows = _build_bilan_rows(participant)
    pdf_path = generate_participant_bilan_pdf(current_app, participant, rows)
    if pdf_path and os.path.exists(pdf_path):
        return send_file(pdf_path, as_attachment=True)
    flash("Impossible de générer le PDF.", "warning")
    return redirect(url_for("statsimpact.stats_pedagogie", participant_id=participant_id))




def _quartier_bucket(name: str | None) -> str:
    """Regroupe les quartiers en 4 grandes catégories + Inconnu."""
    if not name:
        return "Inconnu"
    s = str(name).strip().lower()
    if not s:
        return "Inconnu"
    # règles simples : on privilégie le 'contains' (les noms dans ta DB peuvent varier)
    if "rouher" in s:
        return "Rouher"
    if "bas" in s:
        return "Bas de Creil"
    if "haut" in s or "hauts" in s:
        return "Hauts de Creil"
    # tout le reste
    return "Autres"


def _build_magato_per_atelier_workbook(flt) -> Workbook:
    """Export annuel "Excel" : 1 feuille par atelier (matrice participants x sessions) + bloc de stats riches."""

    # Cloisonnement : un responsable_secteur ne doit exporter que son secteur
    eff_secteur = flt.secteur
    if not can("scope:all_secteurs"):
        eff_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip() or eff_secteur

    # Liste des ateliers dans le périmètre
    aq = AtelierActivite.query.filter(AtelierActivite.is_deleted.is_(False))
    if eff_secteur:
        aq = aq.filter(AtelierActivite.secteur == eff_secteur)
    ateliers = aq.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    wb = Workbook()

    # Synthèse globale
    ws0 = wb.active
    ws0.title = "Synthese"
    ws0.append(["Export annuel : 1 feuille par atelier (matrice) + stats détaillées"])
    ws0.append([])

    ws0.append([
        "Secteur",
        "Atelier",
        "Nb séances prévisionnelles",
        "Nb séances réelles",
        "Nb présences totales",
        "Nb inscrits (uniques)",
        "Capacité d'accueil (réel)",
        "Nb places prévisionnelles",
        "Nb heures prévisionnelles",
        "Nb heures réalisées",
        "Taux d'occupation (%)",
        "Moyenne / séance",
        "Durée de l'activité (jours)",
        "Moyenne d'âge",
        "Bas de Creil",
        "Hauts de Creil",
        "Rouher",
        "Autres",
        "Inconnu",
    ])

    # Agrégats globaux provenance
    global_prov = {"Bas de Creil": 0, "Hauts de Creil": 0, "Rouher": 0, "Autres": 0, "Inconnu": 0}

    for at in ateliers:
        # Sessions de l'atelier dans la période
        sess_q = db.session.query(SessionActivite).filter(SessionActivite.atelier_id == at.id)
        if flt.date_from:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) >= flt.date_from)
        if flt.date_to:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) <= flt.date_to)
        sess_q = sess_q.order_by(
            func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session).asc(),
            SessionActivite.id.asc(),
        )
        sessions = sess_q.all()

        # 1) KPIs par atelier
        sessions_planned = len(sessions)
        sessions_real = sum(1 for s in sessions if (s.statut or "").lower() != "annulee")
        dates = [d for d in [(s.rdv_date or s.date_session) for s in sessions] if d]
        duration_days = (max(dates) - min(dates)).days if dates else None

        # heures + capacités
        planned_hours = 0.0
        real_hours = 0.0
        planned_capacity = 0
        real_capacity = 0
        for s in sessions:
            mins = _session_duration_minutes(s, at) or 0
            h = float(mins) / 60.0 if mins else 0.0
            planned_hours += h
            cap = s.capacite if s.capacite is not None else (getattr(at, "capacite_defaut", 0) or 0)
            planned_capacity += int(cap or 0)
            if (s.statut or "").lower() != "annulee":
                real_hours += h
                real_capacity += int(cap or 0)

        if sessions_planned == 0:
            # atelier sans session dans le filtre -> ligne 0 + pas de feuille matrice
            ws0.append([at.secteur, at.nom, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "", "", 0, 0, 0, 0, 0])
            continue

        session_ids = [s.id for s in sessions]

        pres_rows = db.session.query(PresenceActivite.participant_id, PresenceActivite.session_id)             .filter(PresenceActivite.session_id.in_(session_ids)).all()

        presences_total = len(pres_rows)
        pid_set = sorted({int(pid) for (pid, _sid) in pres_rows if pid is not None})
        inscrits_uniques = len(pid_set)

        # participants (objets) pour âge + provenance
        participants = []
        if pid_set:
            participants = (
                db.session.query(Participant)
                .options(joinedload(Participant.quartier))
                .filter(Participant.id.in_(pid_set))
                .all()
            )

        ages = [p.age for p in participants if getattr(p, "age", None) is not None]
        age_avg = round(sum(ages) / len(ages), 1) if ages else None

        prov = {"Bas de Creil": 0, "Hauts de Creil": 0, "Rouher": 0, "Autres": 0, "Inconnu": 0}
        for p in participants:
            qname = p.quartier.nom if getattr(p, "quartier", None) is not None else None
            b = _quartier_bucket(qname)
            prov[b] = prov.get(b, 0) + 1
        for k in global_prov:
            global_prov[k] += prov.get(k, 0)

        occupation_rate = round((presences_total / real_capacity * 100.0), 1) if real_capacity > 0 else 0.0
        avg_per_session = round((presences_total / sessions_real), 2) if sessions_real > 0 else 0.0

        ws0.append([
            at.secteur,
            at.nom,
            sessions_planned,
            sessions_real,
            presences_total,
            inscrits_uniques,
            real_capacity,
            planned_capacity,
            round(planned_hours, 2),
            round(real_hours, 2),
            occupation_rate,
            avg_per_session,
            duration_days if duration_days is not None else "",
            age_avg if age_avg is not None else "",
            prov.get("Bas de Creil", 0),
            prov.get("Hauts de Creil", 0),
            prov.get("Rouher", 0),
            prov.get("Autres", 0),
            prov.get("Inconnu", 0),
        ])

        # 2) Feuille atelier : bloc stats + matrice
        ws = wb.create_sheet(_safe_sheet_title(f"{at.nom}"))

        ws.append([f"{at.secteur} — {at.nom}"])
        ws.append([])

        ws.append(["Statistiques", "Valeur"])
        ws.append(["Nb séances prévisionnelles", sessions_planned])
        ws.append(["Nb séances réelles", sessions_real])
        ws.append(["Nb présences totales", presences_total])
        ws.append(["Nb inscrits (uniques)", inscrits_uniques])
        ws.append(["Capacité d'accueil (réel)", real_capacity])
        ws.append(["Nb places prévisionnelles", planned_capacity])
        ws.append(["Nb heures prévisionnelles", round(planned_hours, 2)])
        ws.append(["Nb heures réalisées", round(real_hours, 2)])
        ws.append(["Taux d'occupation (%)", occupation_rate])
        ws.append(["Moyenne / séance", avg_per_session])
        ws.append(["Durée de l'activité (jours)", duration_days if duration_days is not None else ""])
        ws.append(["Moyenne d'âge", age_avg if age_avg is not None else ""])

        ws.append([])
        ws.append(["Provenance (quartiers)", "Nb inscrits"])
        ws.append(["Bas de Creil", prov.get("Bas de Creil", 0)])
        ws.append(["Hauts de Creil", prov.get("Hauts de Creil", 0)])
        ws.append(["Rouher", prov.get("Rouher", 0)])
        ws.append(["Autres", prov.get("Autres", 0)])
        ws.append(["Inconnu", prov.get("Inconnu", 0)])

        ws.append([])

        # Matrice participants × sessions
        # (A) Antoine wanted a quick visual identification: âge + ville + quartier à côté du nom/prénom.
        headers = ["Nom", "Prénom", "Âge", "Ville", "Quartier"] + [
            ((d.strftime("%d/%m/%Y")) if (d := (s.rdv_date or s.date_session)) else "Sans date")
            for s in sessions
        ]
        ws.append(headers)

        if not pres_rows or not pid_set:
            # matrice vide structurée
            ws.column_dimensions[get_column_letter(1)].width = 20
            ws.column_dimensions[get_column_letter(2)].width = 18
            ws.column_dimensions[get_column_letter(3)].width = 8
            ws.column_dimensions[get_column_letter(4)].width = 18
            ws.column_dimensions[get_column_letter(5)].width = 22
            for col_idx in range(6, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            continue

        # Participants (objets) pour trier + infos d'identification (âge/ville/quartier)
        parts = (
            db.session.query(Participant)
            .options(joinedload(Participant.quartier))
            .filter(Participant.id.in_(pid_set))
            .order_by(Participant.nom.asc(), Participant.prenom.asc())
            .all()
        )

        sid_index = {int(s.id): idx for idx, s in enumerate(sessions)}
        present = {(int(pid), int(sid)) for (pid, sid) in pres_rows if pid is not None and sid is not None}

        for p in parts:
            pid = int(p.id)
            nom = p.nom or ""
            prenom = p.prenom or ""
            age = getattr(p, "age", None)
            age = age if age is not None else ""
            ville = p.ville or ""
            quartier = p.quartier.nom if getattr(p, "quartier", None) is not None else ""

            row = [nom, prenom, age, ville, quartier] + [""] * len(sessions)
            for sid, idx in sid_index.items():
                if (int(pid), int(sid)) in present:
                    row[5 + idx] = "1"
            ws.append(row)

        # Largeurs
        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.column_dimensions[get_column_letter(2)].width = 18
        ws.column_dimensions[get_column_letter(3)].width = 8
        ws.column_dimensions[get_column_letter(4)].width = 18
        ws.column_dimensions[get_column_letter(5)].width = 22
        for col_idx in range(6, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # Bloc "Provenance globale" en bas de la synthèse
    ws0.append([])
    ws0.append(["Provenance globale (inscrits uniques sur l'ensemble du périmètre)"])
    ws0.append(["Bas de Creil", global_prov.get("Bas de Creil", 0)])
    ws0.append(["Hauts de Creil", global_prov.get("Hauts de Creil", 0)])
    ws0.append(["Rouher", global_prov.get("Rouher", 0)])
    ws0.append(["Autres", global_prov.get("Autres", 0)])
    ws0.append(["Inconnu", global_prov.get("Inconnu", 0)])

    # Largeurs synthèse
    for col in range(1, 20):
        ws0.column_dimensions[get_column_letter(col)].width = 20 if col <= 2 else 18

    return wb


@bp.route("/stats-impact", methods=["GET"])
@login_required
def exports():
    """Page légère dédiée aux exports (pas de calcul de stats lourdes)."""
    if not _can_view():
        abort(403)

    flt = normalize_filters(dict(request.args), user=current_user)

    # Default: current year if no dates
    if not flt.date_from and not flt.date_to:
        today = date.today()
        flt.date_from = date(today.year, 1, 1)
        flt.date_to = date(today.year, 12, 31)

    secteurs = []
    if can("statsimpact:view_all") or can("scope:all_secteurs"):
        secteurs = [
            s[0]
            for s in (
                AtelierActivite.query.with_entities(AtelierActivite.secteur)
                .filter(AtelierActivite.is_deleted.is_(False))
                .distinct()
                .order_by(AtelierActivite.secteur.asc())
                .all()
            )
            if s and s[0]
        ]

    q = AtelierActivite.query.filter(AtelierActivite.is_deleted.is_(False))
    if flt.secteur:
        q = q.filter(AtelierActivite.secteur == flt.secteur)
    ateliers = q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    return render_template(
        "statsimpact/exports.html",
        flt=flt,
        secteurs=secteurs,
        ateliers=ateliers,
        csv_field_groups=CSV_FIELD_GROUPS,
        csv_default_fields=CSV_DEFAULT_FIELDS,
    )

@bp.route("/stats-impact/", methods=["GET"])
@login_required
def exports_slash():
    """Compat: /stats-impact/ -> /stats-impact"""
    if not _can_view():
        abort(403)
    return redirect(url_for("statsimpact.exports", **request.args.to_dict(flat=True)))



@bp.route("/stats-impact/dashboard", methods=["GET", "POST"])
@login_required
def dashboard():
    """Dashboard complet Stats & Impacts (calculs + onglets)."""
    if not _can_view():
        abort(403)

    args = dict(request.args)

    # Robust: normalize_filters supports both dict-style and kwargs-style.
    flt = normalize_filters(args, user=current_user)

    # Default: current year if no dates
    if not flt.date_from and not flt.date_to:
        today = date.today()
        flt.date_from = date(today.year, 1, 1)
        flt.date_to = date(today.year, 12, 31)

    # Pre-compute participants for access control if we need to handle edits.
    participants = compute_participants_stats(flt)

    if request.method == "POST":
        action = request.form.get("action")
        if action == "update_participant":
            try:
                participant_id = int(request.form.get("participant_id", "0"))
            except Exception:
                participant_id = 0

            allowed_ids = {p["id"] for p in participants.get("participants", [])}
            if not participant_id or participant_id not in allowed_ids:
                abort(403)

            participant = Participant.query.get(participant_id)
            if not participant:
                abort(404)

            participant.nom = (request.form.get("nom") or participant.nom or "").strip() or participant.nom
            participant.prenom = (request.form.get("prenom") or participant.prenom or "").strip() or participant.prenom
            participant.ville = (request.form.get("ville") or "").strip() or None
            participant.email = (request.form.get("email") or "").strip() or None
            participant.telephone = (request.form.get("telephone") or "").strip() or None
            participant.genre = (request.form.get("genre") or "").strip() or None
            participant.type_public = (request.form.get("type_public") or participant.type_public or "H").strip().upper()

            dn_raw = request.form.get("date_naissance") or None
            dn = None
            if dn_raw:
                try:
                    dn = date.fromisoformat(dn_raw)
                except Exception:
                    dn = None
            participant.date_naissance = dn

            quartier_id = request.form.get("quartier_id") or None
            try:
                participant.quartier_id = int(quartier_id) if quartier_id else None
            except Exception:
                participant.quartier_id = None

            try:
                from app.extensions import db

                db.session.commit()
                flash("Participant mis à jour.", "success")
            except Exception:
                db.session.rollback()
                flash("Impossible de sauvegarder ce participant.", "danger")

            args_redirect = request.args.to_dict(flat=True)
            args_redirect["tab"] = "participants"
            return redirect(url_for("statsimpact.dashboard", **args_redirect))

        if action == "delete_participant":
            try:
                participant_id = int(request.form.get("participant_id", "0"))
            except Exception:
                participant_id = 0

            allowed_ids = {p["id"] for p in participants.get("participants", [])}
            if not participant_id or participant_id not in allowed_ids:
                abort(403)

            participant = Participant.query.get(participant_id)
            if not participant:
                abort(404)

            # Sécurité secteur: un responsable_secteur ne peut purger un participant
            # que si ce participant n'a des présences que dans SON secteur (ou aucune).
            user_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip()
            if not can("participants:view_all"):
                sectors = (
                    PresenceActivite.query.join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
                    .with_entities(SessionActivite.secteur)
                    .filter(PresenceActivite.participant_id == participant_id)
                    .distinct()
                    .all()
                )
                sectors = {s[0] for s in sectors if s and s[0]}
                # s'il n'a jamais émargé: OK (secteurs = vide)
                if sectors and sectors != {user_secteur}:
                    flash(
                        "Suppression refusée : ce participant a des émargements dans d'autres secteurs.",
                        "danger",
                    )
                    args_redirect = request.args.to_dict(flat=True)
                    args_redirect["tab"] = "participants"
                    return redirect(url_for("statsimpact.dashboard", **args_redirect))

            try:
                from app.extensions import db

                # Supprime d'abord les signatures des présences
                presences = PresenceActivite.query.filter_by(participant_id=participant_id).all()
                for pr in presences:
                    if pr.signature_path:
                        try:
                            if os.path.exists(pr.signature_path):
                                os.remove(pr.signature_path)
                        except Exception:
                            pass
                    db.session.delete(pr)

                db.session.delete(participant)
                db.session.commit()
                flash("Participant supprimé définitivement.", "success")
            except Exception:
                db.session.rollback()
                flash("Impossible de supprimer ce participant.", "danger")

            args_redirect = request.args.to_dict(flat=True)
            args_redirect["tab"] = "participants"
            return redirect(url_for("statsimpact.dashboard", **args_redirect))

    # Refresh computed stats after any potential mutation
    participants = compute_participants_stats(flt)
    stats = compute_volume_activity_stats(flt)
    freq = compute_participation_frequency_stats(flt)
    trans = compute_transversalite_stats(flt)
    demo = compute_demography_stats(flt)
    occupancy = compute_occupancy_stats(flt)

    # Le Magatomatique : calcul uniquement si l'onglet est affiché (sinon on garde la page légère)
    tab = (request.args.get("tab") or "base").strip().lower()
    magato = None
    if tab in ("magato", "magatomatique"):
        participant_q = (request.args.get("participant_q") or "").strip() or None
        view = (request.args.get("magato_view") or "macro").strip().lower()
        try:
            max_sessions = int(request.args.get("max_sessions") or 40)
        except Exception:
            max_sessions = 40
        try:
            max_participants = int(request.args.get("max_participants") or 250)
        except Exception:
            max_participants = 250

        # bornes de sécurité
        max_sessions = max(5, min(max_sessions, 200))
        max_participants = max(20, min(max_participants, 1000))

        magato = compute_magatomatique(
            flt,
            participant_q=participant_q,
            view=view,
            max_sessions=max_sessions,
            max_participants=max_participants,
        )

    participants = compute_participants_stats(flt)

    secteurs = []
    if can("statsimpact:view_all") or can("scope:all_secteurs"):
        secteurs = [
            s[0]
            for s in (
                AtelierActivite.query.with_entities(AtelierActivite.secteur)
                .filter(AtelierActivite.is_deleted.is_(False))
                .distinct()
                .order_by(AtelierActivite.secteur.asc())
                .all()
            )
            if s and s[0]
        ]

    q = AtelierActivite.query.filter(AtelierActivite.is_deleted.is_(False))
    if flt.secteur:
        q = q.filter(AtelierActivite.secteur == flt.secteur)
    ateliers = q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    quartiers = Quartier.query.order_by(Quartier.nom.asc()).all()

    # Années disponibles (pour presets "année") dans le périmètre accessible
    try:
        eff_secteur = flt.secteur
        if not can("scope:all_secteurs"):
            eff_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip() or eff_secteur

        year_expr = func.extract("year", func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session))
        years_q = (
            db.session.query(year_expr.label("y"))
            .select_from(SessionActivite)
            .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
            .filter(AtelierActivite.is_deleted.is_(False))
        )
        if eff_secteur:
            years_q = years_q.filter(AtelierActivite.secteur == eff_secteur)
        years = [int(r.y) for r in years_q.distinct().order_by(year_expr.desc()).all() if r and r.y]
    except Exception:
        years = []

    return render_template(
        ["statsimpact/dashboard.html", "statsimpact_dashboard.html"],
        flt=flt,
        stats=stats,
        freq=freq,
        trans=trans,
        demo=demo,
        secteurs=secteurs,
        ateliers=ateliers,
        occupancy=occupancy,
        participants=participants,
        magato=magato,
        quartiers=quartiers,
        available_years=years,
        csv_field_groups=CSV_FIELD_GROUPS,
        csv_default_fields=CSV_DEFAULT_FIELDS,
    )



@bp.route("/stats-impact/magatomatique.csv", methods=["GET"])
@login_required
def magatomatique_export_csv():
    if not _can_view():
        abort(403)

    flt = normalize_filters(dict(request.args), user=current_user)
    participant_q = (request.args.get("participant_q") or "").strip() or None

    fields = request.args.getlist("fields")
    if not fields:
        fields = list(CSV_DEFAULT_FIELDS)
    fields = [f for f in fields if f in CSV_FIELD_MAP]
    if not fields:
        fields = list(CSV_DEFAULT_FIELDS)

    query = _query_presence_export(flt, participant_q=participant_q)

    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([CSV_FIELD_MAP[f]["label"] for f in fields])

    for presence, participant, session, atelier, quartier in query.all():
        ctx = {
            "presence": presence,
            "participant": participant,
            "session": session,
            "atelier": atelier,
            "quartier": quartier,
        }
        writer.writerow([CSV_FIELD_MAP[f]["getter"](ctx) for f in fields])

    csv_name = "magatomatique_export.csv"
    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename={csv_name}"
    return resp


@bp.route("/stats-impact/magatomatique.xlsx", methods=["GET"])
@login_required
def magatomatique_export():
    if not _can_view():
        abort(403)

    flt = normalize_filters(dict(request.args), user=current_user)

    export_mode = (request.args.get("export_mode") or "flat").strip().lower()
    participant_q = (request.args.get("participant_q") or "").strip() or None
    view = (request.args.get("magato_view") or "macro").strip().lower()
    try:
        max_sessions = int(request.args.get("max_sessions") or 40)
    except Exception:
        max_sessions = 40
    try:
        max_participants = int(request.args.get("max_participants") or 250)
    except Exception:
        max_participants = 250

    # bornes (export raisonnable)
    max_sessions = max(5, min(max_sessions, 400))
    max_participants = max(20, min(max_participants, 5000))

    # Mode "per_atelier" : export annuel 1 feuille = 1 atelier
    if export_mode in ("per_atelier", "per-atelier", "atelier"):
        wb = _build_magato_per_atelier_workbook(flt)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = "magatomatique_par_atelier.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    magato = compute_magatomatique(
        flt,
        participant_q=participant_q,
        view=view,
        max_sessions=max_sessions,
        max_participants=max_participants,
    )

    if magato.get("restricted"):
        abort(403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthese"

    # En-têtes synthèse macro (secteurs)
    ws.append(["Synthèse par secteur"])
    ws.append(["Secteur", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_secteur", []):
        ws.append([r["secteur"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    ws.append([])
    ws.append(["Synthèse par atelier"])
    ws.append(["Secteur", "Atelier", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_atelier", []):
        ws.append([r["secteur"], r["atelier_nom"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    # Feuille participants (si dispo)
    if magato.get("participants"):
        ws2 = wb.create_sheet("Participants")
        ws2.append(["Participants (dans le périmètre filtré)"])
        ws2.append(["Nom", "Prénom", "Ville", "Quartier", "Nb présences", "1ère venue", "Dernière venue"])
        for p in magato["participants"]:
            fd = p.get("first_date")
            ld = p.get("last_date")
            ws2.append([
                p.get("nom",""),
                p.get("prenom",""),
                p.get("ville") or "",
                p.get("quartier") or "",
                int(p.get("nb_presences",0)),
                fd.strftime("%Y-%m-%d") if fd else "",
                ld.strftime("%Y-%m-%d") if ld else "",
            ])

    # Feuille matrice (si view=matrix)
    if magato.get("view") == "matrix" and magato.get("sessions") and magato.get("participants"):
        ws3 = wb.create_sheet("Matrice")
        sessions = magato["sessions"]
        participants = magato["participants"]
        matrix = magato.get("matrix") or {}

        header = ["Nom", "Prénom"] + [f'{s["atelier"]} · {s["label"]}' for s in sessions]
        ws3.append(header)

        for p in participants:
            row = [p.get("nom",""), p.get("prenom","")]
            pid = int(p["id"])
            for s in sessions:
                sid = int(s["id"])
                row.append("1" if matrix.get((pid, sid)) else "")
            ws3.append(row)

        # Ajuste largeur colonnes
        for col_idx in range(1, len(header) + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 16 if col_idx <= 2 else 12

        ws4 = wb.create_sheet("Participations")
        ws4.append(["Nom", "Prénom", "Atelier", "Secteur", "Date session", "ID session"])
        for p in participants:
            pid = int(p["id"])
            for s in sessions:
                sid = int(s["id"])
                if matrix.get((pid, sid)):
                    ws4.append(
                        [
                            p.get("nom", ""),
                            p.get("prenom", ""),
                            s.get("atelier", ""),
                            s.get("secteur", ""),
                            s.get("date").strftime("%Y-%m-%d") if s.get("date") else "",
                            sid,
                        ]
                    )

        for col_idx in range(1, 7):
            ws4.column_dimensions[get_column_letter(col_idx)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "magatomatique.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
